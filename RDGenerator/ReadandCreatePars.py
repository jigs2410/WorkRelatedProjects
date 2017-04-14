# -*- coding: utf-8 -*-
"""
Created on Tue Apr 11 15:38:50 2017

@author: jrathod
"""

# -*- coding: utf-8 -*-
"""
Created on Thu Apr  6 13:31:35 2017

@author: jrathod
"""
import re
rowCount=0
lines=[]
with open("Param_VolumatV2_22b.txt") as fo:
   lines = fo.readlines()
fo.close()

zoneIdentified=0
parIdentified=0
getFields=0

lineincr=0
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'RDParams.xlsx'

ws1 = wb.active
ws1.title = "ParOnly"
#ws1.sheet_properties.tabColor = "1072BA"
#ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
ws1.cell(column=1, row=1, value="ZONE")
ws1.cell(column=2, row=1, value="ZONE NAME")
ws1.cell(column=3, row=1, value="PAR") 
ws1.cell(column=4, row=1, value="PAR NAME")
ws1.cell(column=5, row=1, value="SUBPAR")
ws1.cell(column=6, row=1, value="SUBPAR NAME")
ws1.cell(column=7, row=1, value="VALUE")
ws1.cell(column=8, row=1, value="TYPE")
ws1.cell(column=9, row=1, value="BIT")
ws1.cell(column=10, row=1, value="BIT NAME")

#ws1.cell(column=1, row=0, value="ZONE NAME")

#ws1[0,0]="ZONE"	
#ws1[0,1]="ZONE NAME"	
#ws1[0,2]="PAR"	
#ws1[0,3]="PAR NAME"	
#ws1[0,4]="SUBPAR"
#ws1[0,5]="SUBPAR NAME"
#ws1[0,6]="VALUE"	
#ws1[0,7]="TYPE"	
#ws1[0,8]="BIT"
#ws1[0,9]="BIT NAME"

excelRowCount=2


def writeToWB(rownum,colnum,data,incrRow="No"):
    ws1.cell(column=colnum, row=rownum, value=data)

    global excelRowCount
    
    if (incrRow=="Yes"):
        excelRowCount+=1

for line in lines:
   #print(rowCount)
   matchObj=""
   
   if ("/* Zone" in line or  "/* Notes" in line):
       zoneIdentified=0
       parIdentified=0
       getFields=0
       #print(rowCount,line.strip(' \t\n\r'))
   
   if ("$zone" in line):
       #print(rowCount,"Zone",line.strip(' \t\n\r'))
       matchObj = re.match( r'\$zone (\d*) (\w*) .*', line.strip(' \t\n\r'), re.I)
       #print(rowCount,"Zone",matchObj.group(1),"->",matchObj.group(2))
       print("Zone",matchObj.group(1),"->",matchObj.group(2))
       getFields=0
       parIdentified=0
       zoneIdentified=1
       writeToWB(excelRowCount,1,matchObj.group(1),"No")
       writeToWB(excelRowCount,2,matchObj.group(2),"No")
       
   
   if (zoneIdentified == 1):
       if("$par" in line and "$carearea" in line):
           parIdentified=1
           matchObj = re.match( r'\$par (\d*) (\w*) .*', line, re.I)          #$par (.*) are (.*?) .*', line, re.M|re.I)
           if matchObj:
               #print ("par ID : ", matchObj.group(1))
               #print ("par Name : ", matchObj.group(2))
               getFields=1
               #print(rowCount,"\tPar",matchObj.group(1),"->",matchObj.group(2))
               print("\tPar",matchObj.group(1),"->",matchObj.group(2))
               print("\t\tParams:")
               writeToWB(excelRowCount,3,matchObj.group(1),"No")
               writeToWB(excelRowCount,4,matchObj.group(2),"No")
   
   if(parIdentified==1 and getFields==1):
       #print(rowCount,line)
       if ("Note :" in line):
           parIdentified = 0
           getFields=0
       else: ##extract fields
           if len(line.strip(' \t\n\r')) > 0 and not ("$par" in line):
               words = line.strip(' \t\n\r').split(" ")
               #for word in words:
                   #print(word)
               
               if (words[0].isupper() or words[0][:1]=="$") and (words[0]!="$value"):
                   if (words[0] == "$bit"):
                       #print(rowCount,"\t\t\t\tFlags",line.strip(' \t\n\r'))
                       matchObj = re.match(r'\$bit (\d*) (@\w*).*', line.strip(' \t\n\r'),re.I)
                       if matchObj:
                           #print(rowCount,"\t\t\t",matchObj.group(1),"->",matchObj.group(2),line.strip(' \t\n\r'))
                           #print(rowCount,"\t\t\t",matchObj.group(1),"->",matchObj.group(2))
                           print("\t\t\t\tFlags Bit",matchObj.group(1),"->",matchObj.group(2))
                       #print("\t\t\t\tFlags",line.strip(' \t\n\r'))
                           if(lineincr == 1):
                               writeToWB(excelRowCount-1,9,matchObj.group(1),"No")
                               writeToWB(excelRowCount-1,10,matchObj.group(2),"No")
                               lineincr = 0
                           else:
                               writeToWB(excelRowCount,9,matchObj.group(1),"No")
                               writeToWB(excelRowCount,10,matchObj.group(2),"Yes")
                   else:
                       lineincr=1
                       matchObj = re.match(r'(\w*) (\w*) .*', line.strip(' \t\n\r'),re.I)
                       if matchObj:
                           #print(rowCount,"\t\t\t",matchObj.group(1),"->",matchObj.group(2),line.strip(' \t\n\r'))
                           #print(rowCount,"\t\t\t",matchObj.group(1),"->",matchObj.group(2))
                           print("\t\t\t",matchObj.group(1),"->",matchObj.group(2))
                           writeToWB(excelRowCount,7,matchObj.group(1),"No")
                           writeToWB(excelRowCount,8,matchObj.group(2),"Yes")
                       else:
                           #print("line is->%s"%line)
                           #$struct case
                           matchObj = re.match( r'\$(\w*) (\w*).*', line.strip(' \t\n\r'), re.M|re.I)
                           if matchObj:
                               #print(rowCount,"\t\t\t",matchObj.group(1),"->",matchObj.group(2),line.strip(' \t\n\r'))
                               #print(rowCount,"\t\t\t",matchObj.group(1),"->",matchObj.group(2))
                               print("\t\t\t",matchObj.group(1),"->",matchObj.group(2))
                               writeToWB(excelRowCount,7,matchObj.group(1),"No")
                               writeToWB(excelRowCount,8,matchObj.group(2),"Yes")
                           else:
                               #paramater with round brackets
                               matchObj = re.match( r'([A-Za-z0-9\(\)]*) (\w*).*', line.strip(' \t\n\r'), re.M|re.I)
                               if matchObj:
                                   #print(rowCount,"\t\t\t",matchObj.group(1),"->",matchObj.group(2),line.strip(' \t\n\r'))
                                   #print(rowCount,"\t\t\t",matchObj.group(1),"->",matchObj.group(2))
                                   print("\t\t\t",matchObj.group(1),"->",matchObj.group(2))
                                   writeToWB(excelRowCount,7,matchObj.group(1),"No")
                                   writeToWB(excelRowCount,8,matchObj.group(2),"Yes")
                               else:#paramater with square brackets with unit
                                    matchObj = re.match( r'(\w*\[\w*\])  (\w*).*', line.strip(' \t\n\r'), re.M|re.I)
                                    if matchObj:
                                   #print(rowCount,"\t\t\t",matchObj.group(1),"->",matchObj.group(2),line.strip(' \t\n\r'))
                                       #print(rowCount,"\t\t\t",matchObj.group(1),"->",matchObj.group(2))
                                       print("\t\t\t",matchObj.group(1),"->",matchObj.group(2))
                                       writeToWB(excelRowCount,7,matchObj.group(1),"No")
                                       writeToWB(excelRowCount,8,matchObj.group(2),"Yes")
                                    else:#paramater with square brackets with integer
                                         matchObj = re.match( r'(\w*\[\w*\]) (\w*).*', line.strip(' \t\n\r'), re.M|re.I)
                                         if matchObj:
                                   #print(rowCount,"\t\t\t",matchObj.group(1),"->",matchObj.group(2),line.strip(' \t\n\r'))
                                           #print(rowCount,"\t\t\t",matchObj.group(1),"->",matchObj.group(2))
                                           print("\t\t\t",matchObj.group(1),"->",matchObj.group(2))
                                           writeToWB(excelRowCount,7,matchObj.group(1),"No")
                                           writeToWB(excelRowCount,8,matchObj.group(2),"Yes")
                                         else:
                                             writeToWB(excelRowCount,8,"","Yes")
              
   rowCount+=1
   
   
wb.save(filename = dest_filename)
   