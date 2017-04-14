# -*- coding: utf-8 -*-
"""
Created on Thu Apr 13 11:09:13 2017

@author: jrathod
"""

from glob    import glob
import os
import collections
import re
from collections import OrderedDict

from openpyxl import load_workbook

Excelfilepath  = input("Enter Data Dictionary file location: ")
Excelfilepath=Excelfilepath.replace("\\","\\\\")

#Excelfilepath = r'OneDrugLib-Pump Data Dictionary_Mod1.xlsx'
#Excelfilepath = r'DD.xlsx'

EditDataSheet=""
MappingDataSheet=""
Mapping_ED_List=[]
EditData_ED_List=[]

EDS_Agilia_SP_Std=[] 	
EDS_Agilia_VP_Std=[]	
EDS_Agilia_SP_MC=[]
EDS_Agilia_SP_TIVA=[] 	
EDS_Agilia_SP_PCA=[]	
EDS_Agilia_VP_MC=[]
EDS_EditID=""
EDS_Agilia_SP_Std_Val=""
EDS_Agilia_VP_Std_Val=""
EDS_Agilia_SP_MC_Val=""
EDS_Agilia_SP_TIVA_Val=""
EDS_Agilia_SP_PCA_Val=""	
EDS_Agilia_VP_MC_Val=""

MDS_Agilia_SP_Std=[] 	
MDS_Agilia_VP_Std=[]	
MDS_Agilia_SP_MC=[]
MDS_Agilia_SP_TIVA=[] 	
MDS_Agilia_SP_PCA=[]	
MDS_Agilia_VP_MC=[]
MDS_EditID=""
MDS_Agilia_SP_Std_Val=""
MDS_Agilia_VP_Std_Val=""
MDS_Agilia_SP_MC_Val=""
MDS_Agilia_SP_TIVA_Val=""
MDS_Agilia_SP_PCA_Val=""	
MDS_Agilia_VP_MC_Val=""


#XSDfilepath  = input("Enter XSD's Directory location: ")
##XSDfilepath=XSDfilepath.replace("\\","\\\\")
#XSDfilepath= XSDfilepath+"\\*.xsd"
#
#XSD_File_List = glob(XSDfilepath)

try:
    wb = load_workbook(filename = Excelfilepath, read_only=True)
    #Sheet = input("Enter the name of the sheet to processed: ")
    #EditDataSheet = wb[Sheet]
    EditDataSheet = wb["Edit Data"]
    
    MappingDataSheet = wb["Mapping"]
    #row_count = EditDataSheet.max_row
    rowCount=0
except IOError:
   print ("Error: can\'t find file or read data")
except:
    print("File or Sheet is missing")
else:
    ## process mapping sheet
    #print("Processing Edit Data Sheet")
    
    for row in EditDataSheet.rows:
        rowCount+=1
        EDS_EditID=""
        EDS_Agilia_SP_Std_Val=""
        EDS_Agilia_VP_Std_Val=""
        EDS_Agilia_SP_MC_Val=""
        EDS_Agilia_SP_TIVA_Val=""
        EDS_Agilia_SP_PCA_Val=""	
        EDS_Agilia_VP_MC_Val=""
        if rowCount!=1: #skip header
            EDS_EditID = row[2].value
            if not (EDS_EditID is None):
                if not (EDS_EditID in EditData_ED_List) :
                    EditData_ED_List.append(EDS_EditID)
                    #print(rowCount,"EditData ->",EDS_EditID,"Agilia SP Std ->",row[7].value,"Agilia SP MC ->",row[9].value)
                
                ## the way excel stores data we need to jump to cell on the left is the cell value is None
                EDS_Agilia_SP_Std_Val=row[7].value
                EDS_Agilia_VP_Std_Val=row[8].value
                EDS_Agilia_SP_MC_Val=row[9].value
                EDS_Agilia_SP_TIVA_Val=row[10].value
                EDS_Agilia_SP_PCA_Val=row[11].value
                EDS_Agilia_VP_MC_Val=row[12].value
                
                if EDS_Agilia_VP_MC_Val is None:
                    if EDS_Agilia_SP_PCA_Val is None:
                        if EDS_Agilia_SP_TIVA_Val is None:
                            if EDS_Agilia_SP_MC_Val is None:
                                if EDS_Agilia_VP_Std_Val is None:
                                    if EDS_Agilia_SP_Std_Val is None:
                                        print("EDS_Agilia_SP_Std needs to have value")
                                    else:
                                        EDS_Agilia_VP_MC_Val = EDS_Agilia_SP_PCA_Val = EDS_Agilia_SP_TIVA_Val = EDS_Agilia_SP_MC_Val = EDS_Agilia_VP_Std_Val = EDS_Agilia_SP_Std_Val
                                else:    
                                    EDS_Agilia_VP_MC_Val = EDS_Agilia_SP_PCA_Val = EDS_Agilia_SP_TIVA_Val = EDS_Agilia_SP_MC_Val = EDS_Agilia_VP_Std_Val
                            else:
                                EDS_Agilia_VP_MC_Val = EDS_Agilia_SP_PCA_Val = EDS_Agilia_SP_TIVA_Val = EDS_Agilia_SP_MC_Val
                        else:
                            EDS_Agilia_VP_MC_Val = EDS_Agilia_SP_PCA_Val = EDS_Agilia_SP_TIVA_Val
                    else:
                        EDS_Agilia_VP_MC_Val = EDS_Agilia_SP_PCA_Val
                
                ## if ED not already present  add to respoective pump list
                if not (EDS_EditID in EDS_Agilia_SP_Std) and EDS_Agilia_SP_Std_Val != r'n/a':
                    EDS_Agilia_SP_Std.append(EDS_EditID)
                
                if not (EDS_EditID in EDS_Agilia_VP_Std)and EDS_Agilia_VP_Std_Val != r'n/a':
                    EDS_Agilia_VP_Std.append(EDS_EditID)
                
                if not (EDS_EditID in EDS_Agilia_SP_MC)and EDS_Agilia_SP_MC_Val != r'n/a':
                    EDS_Agilia_SP_MC.append(EDS_EditID)
                    
                if not (EDS_EditID in EDS_Agilia_SP_TIVA)and EDS_Agilia_SP_TIVA_Val != r'n/a':
                    EDS_Agilia_SP_TIVA.append(EDS_EditID)
                
                if not (EDS_EditID in EDS_Agilia_SP_PCA)and EDS_Agilia_SP_PCA_Val != r'n/a':
                    EDS_Agilia_SP_PCA.append(EDS_EditID)
                
                if not (EDS_EditID in EDS_Agilia_VP_MC)and EDS_Agilia_VP_MC_Val != r'n/a':
                    EDS_Agilia_VP_MC.append(EDS_EditID)
                                        
                
    ## not needed but good to sort
    EDS_Agilia_SP_MC = sorted(EDS_Agilia_SP_MC)
    EDS_Agilia_SP_Std = sorted(EDS_Agilia_SP_Std)
    EDS_Agilia_VP_Std = sorted(EDS_Agilia_VP_Std)
    EDS_Agilia_SP_MC = sorted(EDS_Agilia_SP_MC)
    EDS_Agilia_SP_TIVA = sorted(EDS_Agilia_SP_TIVA)
    EDS_Agilia_SP_PCA = sorted(EDS_Agilia_SP_PCA)
    EDS_Agilia_VP_MC = sorted(EDS_Agilia_VP_MC)
    
#    print("EDS_Agilia_SP_MC")
#    for item in EDS_Agilia_SP_MC:
#        print(item)
#    print("EDS_Agilia_SP_Std")
#    for item in EDS_Agilia_SP_Std:
#        print(item)
    
    # unload excel and reload
    for row in MappingDataSheet.rows:
        rowCount+=1
        MDS_EditID=""
        MDS_Agilia_SP_Std_Val=""
        MDS_Agilia_VP_Std_Val=""
        MDS_Agilia_SP_MC_Val=""
        MDS_Agilia_SP_TIVA_Val=""
        MDS_Agilia_SP_PCA_Val=""	
        MDS_Agilia_VP_MC_Val=""
        Algo=""
        Devices=""
        if rowCount!=1: #skip header

            if not(row[14].value is None) and ("ED_" in row[14].value) and not(row[10].value is None):
                Algo = row[14].value.strip(' \t\n\r')
                #AlgoLines = Algo.split('\n')
                regex = r"ED_\d+"
                ED_matches = re.findall(regex, Algo)
#                for match in ED_matches:
#                    print ("Full match: %s" % (match))   
#                
                DeviceList = row[10].value.strip(' \t\n\r').split(',')
                for device in DeviceList:
                    if (device.strip(' \t\n\r')):
                        device = re.sub('\s+',' ', device.strip(' \t\n\r'))
                        if device in ('AG_VP_MC_22','AG_VP_MC_13','AG_VP_MC_14') :
                            for MDS_EditID in ED_matches:
                                  if not (MDS_EditID in MDS_Agilia_VP_MC):
                                      MDS_Agilia_VP_MC.append(MDS_EditID)
                                      
                        if device in ('AG_SP_MC_22') :
                            for MDS_EditID in ED_matches:
                                  if not (MDS_EditID in MDS_Agilia_SP_MC):
                                      MDS_Agilia_SP_MC.append(MDS_EditID)    
                                      
                        if device in ('AG_SP_PCA_30') :
                            for MDS_EditID in ED_matches:
                                  if not (MDS_EditID in MDS_Agilia_SP_PCA):
                                      MDS_Agilia_SP_PCA.append(MDS_EditID)                                         
#                if ("\n" in Algo or "\r" in Algo):
#                    #print("Newline",Algo.strip(' \t\n\r'),"->",Devices)  
#                    AlgoLines = Algo.split('\n')
#                    for line in AlgoLines:
#                        print("Line - >",line)
#                else:
#                    print("No Newline",Algo.strip(' \t\n\r'),"->",Devices)    
                
    for item in MDS_Agilia_SP_MC:
        if not(item in EDS_Agilia_SP_MC):
            print("Present in Mapping Agilia_SP_MC but absent in Edit Data Agilia_SP_MC",item)
   
    for item in EDS_Agilia_SP_MC:
        if not(item in MDS_Agilia_SP_MC):
            print("Present in Edit Data Agilia_SP_MC but absent in Mapping Agilia_SP_MC ",item)
            
    #wb.save(filename = Excelfilepath)                
    wb._archive.close()
    wb = None
    
    
    
    EditData_ED_List=[]
   # wb._archive.close()