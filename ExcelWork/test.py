# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

'''
import re
sum = lambda arg1, arg2: arg1 + arg2;

 

# Now you can call sum as a function
print ("Value of total : ", sum( 10, 20 ))
print ("Value of total : ", sum( 20, 20 ))

str1 = "bool[]"
type = str1

phone = "2004-959-559 # This is Phone Number"

# Delete Python-style comments
num = re.sub(r'#.*$', "", phone)
print ("Phone Num : ", num)


if "String(" in str1:
    type = str1.replace('(',"_").replace(')',"")

if "bool[" in str1:
    type = str1.replace('[]',"_ARR")


#type.replace(')',"")

print(str1)
print(type)



# Open a file
fo = open("foo.txt", "w")
fo.write( "Python is a great language.\nYeah its great!!\n")

# Close opend file
fo.close()


from openpyxl import load_workbook

#wb = load_workbook(filename = 'NewRDData.xlsx')
filepath  = input("Enter mapping file location: ")
#filepath=filepath.replace("\\","\\\\")
print(filepath)
#wb = load_workbook(filename = 'C:\\Users\\jrathod\\Desktop\\WorkDocs\\XMLProjectDocs\\ExcelWork\\NewRDData.xlsx')
try:
    wb = load_workbook(filename = filepath)
    Sheet = input("Enter the name of the sheet to processed: ")
#Sheet1 = wb['Sheet6']
    Sheet1 = wb[Sheet]
    row_count = Sheet1.max_row
    print(row_count)
except IOError:
   print ("Error: can\'t find file or read data")
except:
    print("File or Sheet is missing")

import collections
from collections import OrderedDict

AllDevices=[]
i =0;
print ("Enter device and hit enter. Once done entering device type done")
while i!=1:
    device = input()
    if ((device != "done") or (device != "\"done\"")):
        AllDevices.append(device)
    else:
        i = 1

#AllDevices = ['AG_VP_MC_22','AG_SP_MC_22','AG_SP_MC_22','AG_SP_PCA_30']
 
for Device in AllDevices:
    #exec("dict%d = %s" % (i + 1, {}))
    #exec("Global %s_Zones()"% (Device))
    exec("%s_Zones= OrderedDict()" % (Device))
    exec("%s_Zone= OrderedDict()" % (Device))
    exec("%s_Pars= OrderedDict()" % (Device))
    print(Device)
#AG_VP_MC_22_Zones=OrderedDict()
#AG_SP_MC_22_Zones=OrderedDict()
#AG_SP_PCA_30_Zones=OrderedDict()
#AG_SP_TCI_22_Zones=OrderedDict()    

print(AG_VP_MC_22_Zones)
'''
lines=[]
with open("conf.txt") as fo:
   lines = fo.readlines()
fo.close()
DataDictionary=XSDPath=Sheet=""
AllDevices=[]        
for line in lines:
    print(line)
    if (line[:1] != "#"):
        [key,value]=line.split('=')
        value=value.replace("\"","")
        print("Key is->",key,"  value is ->",value)
        if (key=="DataDictionary"):
            DataDictionary=value
        elif (key=="SheetToProcess"):
            Sheet=value
        elif (key=="XSDDirectory"):
            XSDPath=value
        elif (key=="AllDevices"):
            AllDevices=value.split(",")

print(DataDictionary)
print(XSDPath)
print(Sheet)
print(AllDevices)
'''
from openpyxl import load_workbook
try:
    DataDictionary=DataDictionary.replace("\\","\\\\").strip()
    wb = load_workbook(filename = DataDictionary)
    Sheet1 = wb["Sheet6"]
except IOError as ioex:
    print(ioex)
    print ('errno:', ioex.errno)
    #print('err code:', errno.errorcode[ioex.errno])
    #print ('err message:', os.strerror(ioex.errno))
'''    