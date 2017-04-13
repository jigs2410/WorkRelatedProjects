# -*- coding: utf-8 -*-
"""
Created on Tue Mar 21 11:37:40 2017

@author: jrathod
"""
import re
from openpyxl import load_workbook

wb = load_workbook(filename = 'NewRDData.xlsx')

Sheet = wb['Sheet4']
row_count = Sheet.max_row



DeviceList=['AG_SP_MC_22', 'AG_SP_TCI_22', 'AG_SP_PCA_30']
for row in Sheet.rows:
    #print("'",row[0].value,"'")
    List= str(row[0].value).strip(' \t\n\r').split(',')
    
    for item in List:
        if (item.strip() not in DeviceList) and item.strip()!='None' :
            DeviceList.append(item)
        #item =str(item)
        #item=item.strip(' \t\n\r')
        #eplace(' AG','--AG')
        #item = "".join(item.split()) 
        #item=re.sub(r'\s+', '', item)
        #item=re.sub(r'\t', '', item)
        #print ("'",item.translate(None,string.whitespace),"'")
        #print("'",item.replace(' AG','--AG'),"'")
        #print("'"+item.lstrip()+"'")
    #print(List)
    
#     import re
#>>> re.sub(r'\s+', '', 'strip my spaces')
#'stripmyspaces'
print (DeviceList)