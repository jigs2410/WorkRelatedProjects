# -*- coding: utf-8 -*-
"""
Created on Wed Mar 22 11:54:07 2017

@author: jrathod
"""

# -*- coding: utf-8 -*-
"""
Created on Fri Mar 17 10:21:14 2017

@author: jrathod
"""

from openpyxl import load_workbook

wb = load_workbook(filename = 'NewRDData.xlsx')

Sheet1 = wb['Sheet2']
row_count = Sheet1.max_row

print("rowcount %s" % row_count)

rowCount = 0

Zones={}
Zone={}
Par={}
Pars={}
SubPars={}
SubPar={}
Params={}
Param={}
Flags={}
Flag={}
DeviceTypes={}


Zones.clear()
Zone.clear()
Par.clear()
Pars.clear()
SubPar.clear()
SubPars.clear()
Params.clear()
Param.clear()
Flags.clear()
Flag.clear()
DeviceTypes.clear()

CurrDevices=None
CurrRequired=None
CurrRD_ID=None
CurrFlagsPosition=None
CurrFlagsName=None
CurrParam=None
CurrParamType=None
CurrSubPar=None
CurrSubParName=None
CurrPar=None
CurrParName=None
CurrZone=None
CurrZoneName=None

def helper(row_1 , row_2 , currvalue , parentlist , sublist, sublistkeyname_1 , sublistkeyname_2):
    print(rowCount,parentlist)
    print(rowCount,sublist)
        
    if (not (row_1 is None) and not (row_2 is None)):
        if currvalue!=row_1 and not (currvalue) is None:
            parentlist[currvalue]=sublist.copy()
            sublist.clear()
        currvalue=row_1
        sublist[sublistkeyname_1]=row_1
        sublist[sublistkeyname_2]=row_2

for row in Sheet1.rows:
    rowCount+=1
    if rowCount!=1: #skip header

        #helper (row[8].value,row[9].value,CurrFlagsPosition,**Flags,**Flag,'BitID','BitName')
        # we pass on the flag if flag length is only one
        if (not (row[8].value is None) and not (row[9].value is None)) or len(Flag)>0:
            if CurrFlagsPosition!=row[8].value and not (CurrFlagsPosition) is None:
                Flags[CurrFlagsPosition]=Flag.copy()
                Flag.clear()
            if not (row[8].value is None):
                CurrFlagsPosition=row[8].value
                Flag['BitID']= row[8].value
                Flag['BitName']=row[9].value                             
            
        if (not (row[6].value is None) and not (row[7].value is None)):
            if CurrParam!=row[6].value and not (CurrParam is None):
                if (len(Flags)>0):
                    if not (row[8].value is None):
                        Flags[CurrFlagsPosition]=Flag.copy()
                    Flag.clear()
                    Param['Flags']=Flags.copy()
                    Flags.clear()
                    CurrFlagsPosition=None
                Params[CurrParam]=Param.copy()
                Param.clear()
            CurrParam=row[6].value
            Param['Name']=row[6].value
            Param['Type']=row[7].value
            if not (row[11].value is None):
                Param['Required']=row[11].value
            else:
                Param['Required']='No'
                
            if not (row[12].value is None):
                Param['RD_ID']=row[12].value
            else:
                Param['RD_ID']='RD_0000'
 
   
        if (not (row[4].value is None) and not (row[5].value is None)):
            if CurrSubPar!=row[4].value and not (CurrSubPar is None):
                SubPar['Params'] = Params.copy()
                SubPars[CurrSubPar]=SubPar.copy()
                Params.clear()
                SubPar.clear()
            CurrSubPar=row[4].value
            SubPar['ID']=row[4].value
            SubPar['Name']=row[5].value         

        if (not (row[2].value is None) and not (row[3].value is None)):
            if CurrPar!=row[2].value and not (CurrPar is None):
                if (len(SubPars)<=0):
                    Par['Params'] = Params.copy()
                    Pars[CurrPar]=Par.copy()
                    Params.clear()
                    Par.clear()
                else:
                    Par['SubPars'] = SubPars.copy()
                    Pars[CurrPar]=SubPar.copy()
                    SubPars.clear()
                    Par.clear()
            CurrPar=row[2].value
            Par['ID']=row[2].value
            Par['Name']=row[3].value
        #print(rowCount,"Pars",Pars)
        #print(rowCount,"Par",Par)
        
        
        if (not (row[0].value is None) and not (row[1].value is None)):
            if(CurrZone!=row[0].value and not (CurrZone is None)):
                Zone['Pars']=Pars.copy()
                Zones[CurrZone]=Zone.copy()
                Zone.clear()
                Pars.clear()
            CurrZone=row[0].value
            Zone['ID']=row[0].value
            Zone['Name']=row[1].value
        #print(rowCount,Flags)
        #print(rowCount,Flag)
#print(rowCount,"SubPars",SubPars)
#print(rowCount,"SubPar",SubPar)
#print(rowCount,"Pars",Pars)
#print(rowCount,"Par",Par)

if(rowCount == row_count):
    if(len(Flags)>0):
        Flags[CurrFlagsPosition]=Flag.copy()
        Flag.clear()
        Param['Flags']=Flags.copy()
        Flags.clear()
        
    Params[CurrParam]=Param.copy()  
    
    if (len(SubPars)<=0):
        Par['Params'] = Params.copy()
        Pars[CurrPar]=Par.copy()
    else:
        SubPar['Params'] = Params.copy()
        SubPars[CurrSubPar]=SubPar.copy()
        Par['SubPars'] = SubPars.copy()
        
    Pars[CurrPar]=Par.copy()
    #Par['Params'] = Params.copy()
    #Pars[CurrPar]=Par.copy()
    Zone['Pars']=Pars.copy()
    Zones[CurrZone]=Zone.copy()
    Params.clear()
    Zone.clear()
    Pars.clear()
    SubPars.clear()
    Par.clear()
    Param.clear()
print(Zones)

for zonekey, zonevalue in Zones.items():
    print ("Zone ->",zonekey)
    for zonevaluekey, zonevaluevalue in zonevalue.items():    
        if not type(zonevaluevalue) is dict:
            print("\t%s->%s" % (zonevaluekey,zonevaluevalue))
        else:
            #print("\tPars")
            for parkey, parvalue in zonevaluevalue.items():
                 for parvaluekey, parvaluevalue in parvalue.items():
                     if not type(parvaluevalue) is dict:
                         print("\t\tPar:%s->%s" % (parvaluekey,parvaluevalue))
                     else:
                         if(parvaluekey == 'SubPars'):
                             for subparkey, subparvalue in parvaluevalue.items():
                                 #print("\t\t\tSubPars:%s" % subparkey) 
                                 for subparvaluekey, subparvaluevalue in subparvalue.items():
                                    if not type(subparvaluevalue) is dict:
                                         print("\t\t\tSubPars:%s->%s" % (subparvaluekey,subparvaluevalue))
                                    else:
                                        for paramkey, paramvalue in subparvaluevalue.items():
                                            print("\t\t\t\tParams:")
                                            for paramvaluekey, paramvaluevalue in paramvalue.items():
                                                if not type(paramvaluevalue) is dict:
                                                    print("\t\t\t\t\t%s->%s" % (paramvaluekey,paramvaluevalue))
                                                else:
                                                    print("\t\t\t\t\tFlags:%s" % paramvaluekey)
                                                    for flagkey, flagvalue in paramvaluevalue.items():
                                                        print("\t\t\t\t\t\t%s->%s" % (flagkey,flagvalue))
                         else: 
                             if(parvaluekey == 'Params'):            
                                 for paramkey, paramvalue in parvaluevalue.items():
                                     print("\t\t\tParams:")
                                     for paramvaluekey, paramvaluevalue in paramvalue.items():
                                         if not type(paramvaluevalue) is dict:
                                             print("\t\t\t\t%s->%s" % (paramvaluekey,paramvaluevalue))
                                         else:
                                             print("\t\t\t\tFlags:%s" % paramvaluekey)
                                             for flagkey, flagvalue in paramvaluevalue.items():
                                                 print("\t\t\t\t\t%s->%s" % (flagkey,flagvalue))
                          
             
del Zones