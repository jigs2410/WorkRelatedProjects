# -*- coding: utf-8 -*-
"""
Created on Fri Mar 24 11:29:09 2017

@author: jrathod
"""

# -*- coding: utf-8 -*-
"""
Created on Thu Mar 23 11:21:09 2017

@author: jrathod
"""

# -*- coding: utf-8 -*-
"""
Created on Wed Mar 22 11:54:07 2017

@author: jrathod
"""

# -*- coding: utf-8 -*-
"""
Created on Fri Mar 17 10:21:14 2017

@author: jrathod
"""
import collections
from collections import OrderedDict

from openpyxl import load_workbook


AllDevices = ['AG_VP_MC_22','AG_SP_MC_22','AG_SP_TCI_22','AG_SP_PCA_30']

AG_VP_MC_22_Zones=OrderedDict()
AG_SP_MC_22_Zones=OrderedDict()
AG_SP_PCA_30_Zones=OrderedDict()
AG_SP_TCI_22_Zones=OrderedDict()

AG_VP_MC_22_Zone=OrderedDict()
AG_SP_MC_22_Zone=OrderedDict()
AG_SP_PCA_30_Zone=OrderedDict()
AG_SP_TCI_22_Zone=OrderedDict()

AG_VP_MC_22_Pars=OrderedDict()
AG_SP_MC_22_Pars=OrderedDict()
AG_SP_PCA_30_Pars=OrderedDict()
AG_SP_TCI_22_Pars=OrderedDict()



wb = load_workbook(filename = 'NewRDData.xlsx')

Sheet1 = wb['Sheet6']
row_count = Sheet1.max_row

#print("rowcount %s" % row_count)

rowCount = 0

Zones=OrderedDict()
Zone=OrderedDict()
Par=OrderedDict()
Pars=OrderedDict()
SubPars=OrderedDict()
SubPar=OrderedDict()
Params=OrderedDict()
Param=OrderedDict()
Flags=OrderedDict()
Flag=OrderedDict()
DeviceTypes=OrderedDict()
DeviceType=OrderedDict()


Zones.clear()
Zone.clear()
Par.clear()
Pars.clear()
SubPar.clear()
SubPars.clear()
Params.clear()
Param.clear()
Flags.clear()
Flag.clear()
DeviceTypes=[]
DeviceType=[]

CurrDevices=None
CurrRequired=None
CurrRD_ID=None
CurrFlagsPosition=None
CurrFlagsName=None
CurrParam=None
CurrParamType=None
CurrSubPar=None
CurrSubParName=None
CurrPar=None
CurrParName=None
CurrZone=None
CurrZoneName=None
fileFormatValid = 1
DeviceList=None

def printdevices(Zones, DeviceName):
    
    #Zones = collections.OrderedDict(sorted(Zones.items()))
    
    for zonekey, zonevalue in Zones.items():
        print ("Zone ->",zonekey)
        
        for zonevaluekey, zonevaluevalue in zonevalue.items():
            #print(zonevaluevalue)
            #print (type(zonevaluevalue))
            if not type(zonevaluevalue) is collections.OrderedDict:
                print("\t%s->%s" % (zonevaluekey,zonevaluevalue))
            else:
                #print("\tPars")
                for parkey, parvalue in zonevaluevalue.items():
                     for parvaluekey, parvaluevalue in parvalue.items():
                         if not type(parvaluevalue) is collections.OrderedDict:
                             print("\t\tPar:%s->%s" % (parvaluekey,parvaluevalue))
                         else:
                             
                             if(parvaluekey == 'SubPars'):
                                 #print("SubPars->",parvaluekey)
                                 for subparkey, subparvalue in parvaluevalue.items():
                                     #print("\t\t\tSubPars:%s" % subparkey) 
                                     for subparvaluekey, subparvaluevalue in subparvalue.items():
                                        if not type(subparvaluevalue) is collections.OrderedDict:
                                             print("\t\t\tSubPars:%s->%s" % (subparvaluekey,subparvaluevalue))
                                        else:
                                            for paramkey, paramvalue in subparvaluevalue.items():
                                                print("\t\t\t\tParams:")
                                                for paramvaluekey, paramvaluevalue in paramvalue.items():
                                                    if not type(paramvaluevalue) is collections.OrderedDict:
                                                        print("\t\t\t\t\t%s->%s" % (paramvaluekey,paramvaluevalue))
                                                    else:
                                                        print("\t\t\t\t\tFlags:%s" % paramvaluekey)
                                                        for flagkey, flagvalue in paramvaluevalue.items():
                                                            print("\t\t\t\t\t%s->%s" % (flagkey,flagvalue))
                                                            #for flagvaluekey, flagvaluevalue in flagvalue.items():
                                                                #print("\t\t\t\t\t\t%s->%s" % (flagvaluekey,flagvaluevalue))
                             else:
                                # print("Params->",parvaluekey)
                                 if(parvaluekey == 'Params'):            
                                     for paramkey, paramvalue in parvaluevalue.items():
                                         print("\t\t\tParams:")
                                         for paramvaluekey, paramvaluevalue in paramvalue.items():
                                             if not type(paramvaluevalue) is collections.OrderedDict:
                                                 print("\t\t\t\t%s->%s" % (paramvaluekey,paramvaluevalue))
                                             else:
                                                 print("\t\t\t\tFlags:%s" % paramvaluekey)
                                                 for flagkey, flagvalue in paramvaluevalue.items():
                                                     #for flagvaluekey, flagvaluevalue in flagvalue.items():
                                                         #print("\t\t\t\t\t\t%s->%s" % (flagvaluekey,flagvaluevalue))
                                                    print("\t\t\t\t\t%s->%s" % (flagkey,flagvalue))
                          
                            

def helper(row_1 , row_2 , currvalue , parentlist , sublist, sublistkeyname_1 , sublistkeyname_2):
    print(rowCount,parentlist)
    print(rowCount,sublist)
        
    if (not (row_1 is None) and not (row_2 is None)):
        if currvalue!=row_1 and not (currvalue) is None:
            parentlist[currvalue]=sublist.copy()
            sublist.clear()
        currvalue=row_1
        sublist[sublistkeyname_1]=row_1
        sublist[sublistkeyname_2]=row_2

for row in Sheet1.rows:
    rowCount+=1
    if rowCount!=1: #skip header
        if not (row[10].value is None):
            if (rowCount == 2):
                CurrDevices=str(row[10].value).replace(" ", "").strip(' \t\n\r')
            else:    
                CurrDevices=DeviceList
            DeviceList=str(row[10].value).replace(" ", "").split(',')
        else:
            print("Incorrect File Format Device Type not present on row number %s" %rowCount)
            fileFormatValid = 0
            break
        
        
        #helper (row[8].value,row[9].value,CurrFlagsPosition,**Flags,**Flag,'BitID','BitName')
        # we pass on the flag if flag length is only one
        if (not (row[8].value is None) and not (row[9].value is None)) or len(Flag)>0:
            if CurrFlagsPosition!=row[8].value and not (CurrFlagsPosition) is None:
                Flags[CurrFlagsPosition]=Flag.copy()
                Flag.clear()
            if not (row[8].value is None):
                CurrFlagsPosition=row[8].value
                Flag['BitID']= row[8].value
                Flag['BitName']=row[9].value                             
            
        if (not (row[6].value is None) and not (row[7].value is None)):
            if CurrParam!=row[6].value and not (CurrParam is None):
                if (len(Flags)>0):
                    if not (row[8].value is None):
                        Flags[CurrFlagsPosition]=Flag.copy()
                    Flag.clear()
                    Param['Flags']=Flags.copy()
                    Flags.clear()
                    CurrFlagsPosition=None
                Params[CurrParam]=Param.copy()
                Param.clear()
            CurrParam=row[6].value
            Param['Name']=row[6].value
            Param['Type']=row[7].value
            if not (row[11].value is None):
                Param['Required']=row[11].value
            else:
                Param['Required']='No'
                
            if not (row[12].value is None):
                Param['RD_ID']=row[12].value
            else:
                Param['RD_ID']='RD_0000'
 
        #print(rowCount,"Param",Param)
        #print(rowCount,"Params",Params) 
        #print()
        if (not (row[4].value is None) and not (row[5].value is None)):
            if CurrSubPar!=row[4].value and not (CurrSubPar is None):
                SubPar['Params'] = Params.copy()
                SubPars[CurrSubPar]=SubPar.copy()
                Params.clear()
                SubPar.clear()
            CurrSubPar=row[4].value
            SubPar['ID']=row[4].value
            SubPar['Name']=row[5].value   
                  
        #print(rowCount,"SubPar",SubPar)
        #print(rowCount,"SubPars",SubPars)
        #print()        
        if (not (row[2].value is None) and not (row[3].value is None)):
            if (CurrPar!=row[2].value and not (CurrPar is None)) or (CurrDevices!=DeviceList and rowCount!=2):
                if (len(SubPars)<=0):
                    Par['Params'] = Params.copy()
                    
                    if 'AG_VP_MC_22' in CurrDevices:
                        AG_VP_MC_22_Pars[CurrPar]=Par.copy()
                        
                    if 'AG_SP_MC_22' in CurrDevices:
                        AG_SP_MC_22_Pars[CurrPar]=Par.copy()
                    
                    if 'AG_SP_PCA_30' in CurrDevices:
                        AG_SP_PCA_30_Pars[CurrPar]=Par.copy()
                        
                    if 'AG_SP_TCI_22' in CurrDevices:
                        AG_SP_TCI_22_Pars[CurrPar]=Par.copy()
                        
                    #print(rowCount,"Par",Par)
                    #print(rowCount,"Pars",Pars)      
                    Pars[CurrPar]=Par.copy()
                    Params.clear()
                    Par.clear()
                    #print(rowCount,"Par",Par)
                    #print(rowCount,"Pars",Pars)  
                else:
                    Par['SubPars'] = SubPars.copy()
                    Pars[CurrPar]=SubPar.copy()
                    SubPars.clear()
                    Par.clear()
            CurrPar=row[2].value
            Par['ID']=row[2].value
            Par['Name']=row[3].value
        #print(rowCount,"Pars",Pars)
        #print(rowCount,"Par",Par)
        #print()
        #print(rowCount,"Par",Par)
        #print(rowCount,"Pars",Pars)
        
        if (not (row[0].value is None) and not (row[1].value is None)):
            if(CurrZone!=row[0].value and not (CurrZone is None)):
                
                if len(AG_VP_MC_22_Pars) > 0:
                    AG_VP_MC_22_Zone = Zone.copy()
                    AG_VP_MC_22_Zone['Pars'] = AG_VP_MC_22_Pars.copy()
                    AG_VP_MC_22_Pars.clear()
                    
                if len(AG_VP_MC_22_Zone) > 0:
                    AG_VP_MC_22_Zones[CurrZone]=AG_VP_MC_22_Zone.copy()
                    AG_VP_MC_22_Zone.clear()
                    
                if len(AG_SP_MC_22_Pars) > 0:
                    AG_SP_MC_22_Zone = Zone.copy()
                    AG_SP_MC_22_Zone['Pars'] = AG_SP_MC_22_Pars.copy()
                    AG_SP_MC_22_Pars.clear()
                    
                if len(AG_SP_MC_22_Zone) > 0:
                    AG_SP_MC_22_Zones[CurrZone]=AG_SP_MC_22_Zone.copy()
                    AG_SP_MC_22_Zone.clear()
                
                
                
                if len(AG_SP_TCI_22_Pars) > 0:
                    AG_SP_TCI_22_Zone['Pars'] = AG_SP_TCI_22_Pars.copy()
                    AG_SP_TCI_22_Pars.clear()
                    
                if len(AG_SP_TCI_22_Zone) > 0:
                    AG_SP_TCI_22_Zones[CurrZone]=AG_SP_TCI_22_Zone.copy()
                    AG_SP_TCI_22_Zone.clear()



                if len(AG_SP_PCA_30_Pars) > 0:
                    AG_SP_PCA_30_Zone['Pars'] = AG_SP_PCA_30_Pars.copy()
                    AG_SP_PCA_30_Pars.clear()
                    
                if len(AG_SP_PCA_30_Zone) > 0:
                    AG_SP_PCA_30_Zones[CurrZone]=AG_SP_PCA_30_Zone.copy()
                    AG_SP_PCA_30_Zone.clear()                    
                
                Zone['Pars']=Pars.copy()
                Zones[CurrZone]=Zone.copy()
                Zone.clear()
                Pars.clear()
            CurrZone=row[0].value
            Zone['ID']=row[0].value
            Zone['Name']=row[1].value
        #print(rowCount,Flags)
        #print(rowCount,Flag)
#print(rowCount,"SubPars",SubPars)
#print(rowCount,"SubPar",SubPar)


if(rowCount == row_count):
   
    
    
    if(len(Flags)>0):
        Flags[CurrFlagsPosition]=Flag.copy()
        Flag.clear()
        Param['Flags']=Flags.copy()
        Flags.clear()
        
    Params[CurrParam]=Param.copy()  
    
    if (len(SubPars)<=0):
        Par['Params'] = Params.copy()
        Pars[CurrPar]=Par.copy()
    else:
        SubPar['Params'] = Params.copy()
        SubPars[CurrSubPar]=SubPar.copy()
        Par['SubPars'] = SubPars.copy()
        
    Pars[CurrPar]=Par.copy()
    
    
    
    
    #Pars[CurrPar]=Par.copy()
    #Par['Params'] = Params.copy()
    #Pars[CurrPar]=Par.copy()
    Zone['Pars']=Pars.copy()
    Zones[CurrZone]=Zone.copy()

    if 'AG_VP_MC_22' in CurrDevices:
        AG_VP_MC_22_Pars[CurrPar]=Par.copy()

    if len(AG_VP_MC_22_Pars) > 0:
        AG_VP_MC_22_Zone = Zone.copy()
        AG_VP_MC_22_Zone['Pars'] = AG_VP_MC_22_Pars.copy()
        AG_VP_MC_22_Pars.clear()
    
    if len(AG_VP_MC_22_Zone) > 0:
        AG_VP_MC_22_Zones[CurrZone]=AG_VP_MC_22_Zone.copy()
        AG_VP_MC_22_Zone.clear()
    
    if 'AG_SP_MC_22' in CurrDevices:
        AG_SP_MC_22_Pars[CurrPar]=Par.copy()
  
    if len(AG_SP_MC_22_Pars) > 0:
        AG_SP_MC_22_Zone = Zone.copy()
        AG_SP_MC_22_Zone['Pars'] = AG_SP_MC_22_Pars.copy()
        AG_SP_MC_22_Pars.clear()

    if len(AG_SP_MC_22_Zone) > 0:
        AG_SP_MC_22_Zones[CurrZone]=AG_SP_MC_22_Zone.copy()
        AG_SP_MC_22_Zone.clear()
  
    if 'AG_SP_TCI_22' in CurrDevices:
        AG_SP_TCI_22_Pars[CurrPar]=Par.copy()
  
    if len(AG_SP_TCI_22_Pars) > 0:
        AG_SP_TCI_22_Zone = Zone.copy()
        AG_SP_TCI_22_Zone['Pars'] = AG_SP_TCI_22_Pars.copy()
        AG_SP_TCI_22_Pars.clear()

    if len(AG_SP_TCI_22_Zone) > 0:
        AG_SP_TCI_22_Zones[CurrZone]=AG_SP_TCI_22_Zone.copy()
        AG_SP_TCI_22_Zone.clear()
             
    if 'AG_SP_PCA_30' in CurrDevices:
        AG_SP_PCA_30_Pars[CurrPar]=Par.copy()
  
    if len(AG_SP_PCA_30_Pars) > 0:
        AG_SP_PCA_30_Zone = Zone.copy()
        AG_SP_PCA_30_Zone['Pars'] = AG_SP_PCA_30_Pars.copy()
        AG_SP_PCA_30_Pars.clear()

    if len(AG_SP_PCA_30_Zone) > 0:
        AG_SP_PCA_30_Zones[CurrZone]=AG_SP_PCA_30_Zone.copy()
        AG_SP_PCA_30_Zone.clear()
        
#clear    
Params.clear()
Zone.clear()
Pars.clear()
SubPars.clear()
Par.clear()
Param.clear()

if (fileFormatValid == 1):    
    print(Zones)
    #print(AG_VP_MC_22_Zones)
    #print(AG_SP_MC_22_Zones).
    

    #print("###########################  Zones  ####################################")
    #print ("Value : %s" %  Zones.keys())
    #print ("Value ", Zones[8])
    #printdevicesOrder(Zones,'Zones')
    printdevices(Zones,'Zones')
    #print("###########################  AG_VP_MC_22_Zones  ####################################")
    #printdevices(AG_VP_MC_22_Zones,'AG_VP_MC_22')
    #print("###########################  AG_SP_MC_22_Zones  ####################################")
    #printdevices(AG_SP_MC_22_Zones,'AG_SP_MC_22')
    #print("###########################  AG_SP_PCA_30_Zones  ####################################")
    #printdevices(AG_SP_PCA_30_Zones,'AG_SP_PCA_30')
    #print("###########################  AG_SP_TCI_22_Zones  ####################################")
    #printdevices(AG_SP_TCI_22_Zones,'AG_SP_TCI_22')
    '''
    for zonekey, zonevalue in Zones.items():
        print ("Zone ->",zonekey)
        for zonevaluekey, zonevaluevalue in zonevalue.items():    
            if not type(zonevaluevalue) is collections.OrderedDict:
                print("\t%s->%s" % (zonevaluekey,zonevaluevalue))
            else:
                #print("\tPars")
                for parkey, parvalue in zonevaluevalue.items():
                     for parvaluekey, parvaluevalue in parvalue.items():
                         if not type(parvaluevalue) is collections.OrderedDict:
                             print("\t\tPar:%s->%s" % (parvaluekey,parvaluevalue))
                         else:
                             if(parvaluekey == 'SubPars'):
                                 for subparkey, subparvalue in parvaluevalue.items():
                                     #print("\t\t\tSubPars:%s" % subparkey) 
                                     for subparvaluekey, subparvaluevalue in subparvalue.items():
                                        if not type(subparvaluevalue) is collections.OrderedDict:
                                             print("\t\t\tSubPars:%s->%s" % (subparvaluekey,subparvaluevalue))
                                        else:
                                            for paramkey, paramvalue in subparvaluevalue.items():
                                                print("\t\t\t\tParams:")
                                                for paramvaluekey, paramvaluevalue in paramvalue.items():
                                                    if not type(paramvaluevalue) is collections.OrderedDict:
                                                        print("\t\t\t\t\t%s->%s" % (paramvaluekey,paramvaluevalue))
                                                    else:
                                                        print("\t\t\t\t\tFlags:%s" % paramvaluekey)
                                                        for flagkey, flagvalue in paramvaluevalue.items():
                                                            print("\t\t\t\t\t\t%s->%s" % (flagkey,flagvalue))
                             else: 
                                 if(parvaluekey == 'Params'):            
                                     for paramkey, paramvalue in parvaluevalue.items():
                                         print("\t\t\tParams:")
                                         for paramvaluekey, paramvaluevalue in paramvalue.items():
                                             if not type(paramvaluevalue) is collections.OrderedDict:
                                                 print("\t\t\t\t%s->%s" % (paramvaluekey,paramvaluevalue))
                                             else:
                                                 print("\t\t\t\tFlags:%s" % paramvaluekey)
                                                 for flagkey, flagvalue in paramvaluevalue.items():
                                                     print("\t\t\t\t\t%s->%s" % (flagkey,flagvalue))
                          
    '''
del Zones
del AG_VP_MC_22_Zones
del AG_SP_MC_22_Zones