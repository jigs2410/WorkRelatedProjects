# -*- coding: utf-8 -*-
"""
Created on Mon Apr  3 18:23:19 2017

@author: jrathod
"""

"""
Created on Fri Mar 17 10:21:14 2017

@author: jrathod
"""
import collections
from collections import OrderedDict

from openpyxl import load_workbook

from  printXSDToFile import printToFile

try:
    DataDictionary=XSDPath=Sheet=""
    AllDevices=[]      
    lines=[]
    with open("conf.txt") as fo: #reading configuration files
       lines = fo.readlines()
    fo.close()
      
    for line in lines:
        [key,value]=line.split('=')
        value=value.replace("\"","").strip()
        #print("Key is->",key,"  value is ->",value)
        if (key=="DataDictionary"):
            DataDictionary=value
        elif (key=="SheetToProcess"):
            Sheet=value
        elif (key=="XSDDirectory"):
            XSDPath=value
        elif (key=="AllDevices"):
            AllDevices=value.replace("\"","").strip().split(",")
    
    DataDictionary=DataDictionary.replace("\\","\\\\").strip()
    wb = load_workbook(filename = DataDictionary)
    Sheet1 = wb[Sheet]
    row_count = Sheet1.max_row
    
except IOError:
   print ("Error: can\'t find file or read data")
except:
    print("File or Sheet is missing")
else:    
    ###define all the devicelists here
    #AllDevices = ['AG_VP_MC_22','AG_SP_MC_22','AG_SP_TCI_22','AG_SP_PCA_30','AG_SP_PCA_31']
    
    #generate structures on fly
    for Device in AllDevices:
        exec("%s_Zones= OrderedDict()" % (Device))
        exec("%s_Zone= OrderedDict()" % (Device))
        exec("%s_Pars= OrderedDict()" % (Device))
    
    row_count = Sheet1.max_row
    rowCount = 0
    
    #initialize
    Zones=OrderedDict()
    Zone=OrderedDict()
    Par=OrderedDict()
    Pars=OrderedDict()
    SubPars=OrderedDict()
    SubPar=OrderedDict()
    Params=OrderedDict()
    Param=OrderedDict()
    Flags=OrderedDict()
    Flag=OrderedDict()
    DeviceTypes=OrderedDict()
    DeviceType=OrderedDict()
    
    
    Zones.clear()
    Zone.clear()
    Par.clear()
    Pars.clear()
    SubPar.clear()
    SubPars.clear()
    Params.clear()
    Param.clear()
    Flags.clear()
    Flag.clear()
    DeviceTypes=[]
    DeviceType=[]
    
    CurrDevices=None
    CurrRequired=None
    CurrRD_ID=None
    CurrFlagsPosition=None
    CurrFlagsName=None
    CurrParam=None
    CurrParamType=None
    CurrSubPar=None
    CurrSubParName=None
    CurrPar=None
    CurrParName=None
    CurrZone=None
    CurrZoneName=None
    fileFormatValid = 1
    DeviceList=None
   
    #print treebased structural representation
    # if pars doesnt have subpars than params are the child or else params are child of subpars
    def printZoneTree (Zones, DeviceName):
       for zonekey, zonevalue in Zones.items():
            for zonevaluekey, zonevaluevalue in zonevalue.items():
                if not type(zonevaluevalue) is collections.OrderedDict:
                    print("\t%s->%s" % (zonevaluekey,zonevaluevalue))
                else:
                    for parkey, parvalue in zonevaluevalue.items():
                         for parvaluekey, parvaluevalue in parvalue.items():
                             if not type(parvaluevalue) is collections.OrderedDict:
                                 print("\t\tPar:%s->%s" % (parvaluekey,parvaluevalue))
                             else:
                                 if(parvaluekey == 'SubPars'):
                                     for subparkey, subparvalue in parvaluevalue.items():
                                         for subparvaluekey, subparvaluevalue in subparvalue.items():
                                            if not type(subparvaluevalue) is collections.OrderedDict:
                                                 print("\t\t\tSubPars:%s->%s" % (subparvaluekey,subparvaluevalue))
                                            else:
                                                for paramkey, paramvalue in subparvaluevalue.items():
                                                    print("\t\t\t\tParams:")
                                                    for paramvaluekey, paramvaluevalue in paramvalue.items():
                                                        if not type(paramvaluevalue) is collections.OrderedDict:
                                                            print("\t\t\t\t\t%s->%s" % (paramvaluekey,paramvaluevalue))
                                                        else:
                                                            print("\t\t\t\t\tFlags:%s" % paramvaluekey)
                                                            for flagkey, flagvalue in paramvaluevalue.items():
                                                                print("\t\t\t\t\t%s->%s" % (flagkey,flagvalue))
                                 else:
                                    if(parvaluekey == 'Params'):            
                                         for paramkey, paramvalue in parvaluevalue.items():
                                             print("\t\t\tParams:")
                                             for paramvaluekey, paramvaluevalue in paramvalue.items():
                                                 if not type(paramvaluevalue) is collections.OrderedDict:
                                                     print("\t\t\t\t%s->%s" % (paramvaluekey,paramvaluevalue))
                                                 else:
                                                     print("\t\t\t\tFlags:%s" % paramvaluekey)
                                                     for flagkey, flagvalue in paramvaluevalue.items():
                                                        print("\t\t\t\t\t%s->%s" % (flagkey,flagvalue))
    
    # start reading the file and create structure
    for row in Sheet1.rows:
        rowCount+=1
        if rowCount!=1: #skip header
        #general logic is when we read next line previous states are compared and if zone value changes previous
        #data is pushed to Zones collection
        #Similarly if par value has changed that previous par value along with its child is pushed to pars collection
        
            if not (row[10].value is None): #make sure devices are listed.
                if (rowCount == 2):
                    CurrDevices=str(row[10].value).replace(" ", "").strip(' \t\n\r')
                else:    
                    CurrDevices=DeviceList
                DeviceList=str(row[10].value).replace(" ", "").split(',')
            else:
                print("Incorrect File Format Device Type not present on row number %s" %rowCount)
                fileFormatValid = 0
                break
            
            # we pass on the flag if flag length is only one
            if (not (row[8].value is None) and not (row[9].value is None)) or len(Flag)>0:
                if CurrFlagsPosition!=row[8].value and not (CurrFlagsPosition) is None:
                    Flags[CurrFlagsPosition]=Flag.copy()
                    Flag.clear()
                if not (row[8].value is None):
                    CurrFlagsPosition=row[8].value
                    Flag['BitID']= row[8].value
                    Flag['BitName']=row[9].value                             
            
            # for now if we dont have RD value we are defaulting its to RD_0000
            # Same for required column if there is no value its defaulted to No
            if (not (row[6].value is None) and not (row[7].value is None)):
                if CurrParam!=row[6].value and not (CurrParam is None):
                    if (len(Flags)>0):
                        if not (row[8].value is None):
                            Flags[CurrFlagsPosition]=Flag.copy()
                        Flag.clear()
                        Param['Flags']=Flags.copy()
                        Flags.clear()
                        CurrFlagsPosition=None
                    Params[CurrParam]=Param.copy()
                    Param.clear()
                CurrParam=row[6].value
                Param['Name']=row[6].value
                Param['Type']=row[7].value
                if not (row[11].value is None):
                    Param['Required']=row[11].value
                else:
                    Param['Required']='No'
                    
                if not (row[12].value is None):
                    Param['RD_ID']=row[12].value
                else:
                    Param['RD_ID']='RD_0000'
    
            if (not (row[4].value is None) and not (row[5].value is None)):
                if CurrSubPar!=row[4].value and not (CurrSubPar is None):
                    SubPar['Params'] = Params.copy()
                    SubPars[CurrSubPar]=SubPar.copy()
                    Params.clear()
                    SubPar.clear()
                CurrSubPar=row[4].value
                SubPar['ID']=row[4].value
                SubPar['Name']=row[5].value   
            
            # depending on existence of subpars params are linked to pars or subpars
            if (not (row[2].value is None) and not (row[3].value is None)):
                if (CurrPar!=row[2].value and not (CurrPar is None)) or (CurrDevices!=DeviceList and rowCount!=2):
                    if (len(SubPars)<=0):
                        Par['Params'] = Params.copy()
                        
                        for Device in AllDevices:
                            if Device in CurrDevices:
                                exec("%s_Pars[CurrPar]=Par.copy()"%Device)
    
                        Pars[CurrPar]=Par.copy()
                        Params.clear()
                        Par.clear()
                    else:
                        Par['SubPars'] = SubPars.copy()
                        Pars[CurrPar]=SubPar.copy()
                        SubPars.clear()
                        Par.clear()
                CurrPar=row[2].value
                Par['ID']=row[2].value
                Par['Name']=row[3].value
            
            # we collect for alll difference devices mentioned in conf.txt and 
            # maintain individual structure
            if (not (row[0].value is None) and not (row[1].value is None)):
                if(CurrZone!=row[0].value and not (CurrZone is None)):
                    
                    def forallDevices(Device_Pars, Device_Zone, Device_Zones):
                        if len(Device_Pars) > 0:
                            Device_Zone = Zone.copy()
                            Device_Zone['Pars'] = Device_Pars.copy()
                            Device_Pars.clear()
                        
                        if len(Device_Zone) > 0:
                            Device_Zones[CurrZone]=Device_Zone.copy()
                            Device_Zone.clear()
                        
                    for Device in AllDevices:
                        exec("forallDevices(%s_Pars,%s_Zone,%s_Zones)"% (Device,Device,Device))
    
                    Zone['Pars']=Pars.copy()
                    Zones[CurrZone]=Zone.copy()
                    Zone.clear()
                    Pars.clear()
                CurrZone=row[0].value
                Zone['ID']=row[0].value
                Zone['Name']=row[1].value
    
    # for the last row whatever is accumulated gets assgined and cleared
    if(rowCount == row_count):
        if(len(Flags)>0):
            Flags[CurrFlagsPosition]=Flag.copy()
            Flag.clear()
            Param['Flags']=Flags.copy()
            Flags.clear()
            
        Params[CurrParam]=Param.copy()  
        
        if (len(SubPars)<=0):
            Par['Params'] = Params.copy()
            Pars[CurrPar]=Par.copy()
        else:
            SubPar['Params'] = Params.copy()
            SubPars[CurrSubPar]=SubPar.copy()
            Par['SubPars'] = SubPars.copy()
            
        Pars[CurrPar]=Par.copy()
        Zone['Pars']=Pars.copy()
        Zones[CurrZone]=Zone.copy()
        
        def forallDevicesLastLine(Device,Device_Pars, Device_Zone, Device_Zones):
            if Device in CurrDevices:
                Device_Pars[CurrPar]=Par.copy()
    
            if len(Device_Pars) > 0:
                Device_Zone = Zone.copy()
                Device_Zone['Pars'] = Device_Pars.copy()
                Device_Pars.clear()
        
            if len(Device_Zone) > 0:
                Device_Zones[CurrZone]=Device_Zone.copy()
                Device_Zone.clear()
        
        #dynamic function        
        for Device in AllDevices:
            exec("forallDevicesLastLine('%s',%s_Pars,%s_Zone,%s_Zones)"% (Device,Device,Device,Device))            
    #clear    
    Params.clear()
    Zone.clear()
    Pars.clear()
    SubPars.clear()
    Par.clear()
    Param.clear()
    
    #create xsd's at targetdir
    if (fileFormatValid == 1):    
        printToFile(Zones,'Zones',XSDPath)
        #printZoneTree(Zones,'Zones')
        for Device in AllDevices:
            exec("printToFile(%s_Zones,'%s',r'%s')" % (Device,Device,XSDPath))
        