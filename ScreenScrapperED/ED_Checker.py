# -*- coding: utf-8 -*-
"""
Created on Fri Mar 31 09:26:30 2017

@author: jrathod
"""
from glob    import glob
import os
import collections
import re
from collections import OrderedDict

from openpyxl import load_workbook

Excelfilepath  = input("Enter Data Dictionary file location: ")
Excelfilepath=Excelfilepath.replace("\\","\\\\")

#Excelfilepath = r'OneDrugLib-Pump Data Dictionary.xlsx'


Sheet1=""
Excel_ED_List=[]
XSD_ED_Dict={}

XSDfilepath  = input("Enter XSD's Directory location: ")
#XSDfilepath=XSDfilepath.replace("\\","\\\\")
XSDfilepath= XSDfilepath+"\\*.xsd"

XSD_File_List = glob(XSDfilepath)

try:
    wb = load_workbook(filename = Excelfilepath)
    #Sheet = input("Enter the name of the sheet to processed: ")
    #Sheet1 = wb[Sheet]
    Sheet1 = wb["Edit Data"]
    row_count = Sheet1.max_row
    rowCount=0
except IOError:
   print ("Error: can\'t find file or read data")
except:
    print("File or Sheet is missing")
else:
    print("row_counts is ", row_count)
    
    for row in Sheet1.rows:
        rowCount+=1
        if rowCount!=1: #skip header
            if not (row[2].value is None):
                if not (row[2].value in Excel_ED_List) :
                    Excel_ED_List.append(row[2].value)
                else:
                    print(row[2].value,"Duplicate entries in Excel")
    
    #for eds in Excel_ED_List:
        #print(eds)
  
    for xsd in XSD_File_List:
        lines=[]
        with open(xsd) as fo:
            lines = fo.readlines()
        fo.close()
        
        for line in lines:
            if "ED_" in line:
                startindex = line.index("ED_")
                endindex = line.index("\"",startindex)
                ed = line[startindex:endindex]
                #print(line," ",ed," ",len(ed))
                if not (ed in XSD_ED_Dict):
                    #XSD_ED_Dict.append(ed)
                    XSD_ED_Dict[ed]=xsd
        #print(lines)
        #print(xsd)
    
    #for eds in XSD_ED_Dict:
        #print(eds)
        
    #now compare and print
    print("ED Absent in Edit Data Dictionary but present in XSD ")
    for eds in XSD_ED_Dict:
        if not (eds in Excel_ED_List):
            print(eds,"->",XSD_ED_Dict[eds] )
    print("")            
    print("ED not present in any XSD's ")
    for eds in Excel_ED_List:
        if not (eds in XSD_ED_Dict):
            print(eds )