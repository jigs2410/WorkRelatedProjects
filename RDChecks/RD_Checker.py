# -*- coding: utf-8 -*-
"""
Created on Thu Apr 13 11:09:13 2017

@author: jrathod
"""

from glob    import glob
import os
import collections
import re
from collections import OrderedDict

from openpyxl import load_workbook

Excelfilepath  = input("Enter Data Dictionary file location: ")
Excelfilepath=Excelfilepath.replace("\\","\\\\")

#Excelfilepath = r'OneDrugLib-Pump Data Dictionary.xlsx'


MappingSheet=""
Mapping_RD_List=[]
Agilia_RD_List=[]

#XSDfilepath  = input("Enter XSD's Directory location: ")
##XSDfilepath=XSDfilepath.replace("\\","\\\\")
#XSDfilepath= XSDfilepath+"\\*.xsd"
#
#XSD_File_List = glob(XSDfilepath)

try:
    wb = load_workbook(filename = Excelfilepath, read_only=True)
    #Sheet = input("Enter the name of the sheet to processed: ")
    #MappingSheet = wb[Sheet]
    MappingSheet = wb["Mapping"]
    
    #row_count = MappingSheet.max_row
    rowCount=0
except IOError:
   print ("Error: can\'t find file or read data")
except:
    print("File or Sheet is missing")
else:
    ## process mapping sheet
    print("Processing Mapping Sheet")
    for row in MappingSheet.rows:
        rowCount+=1
        if rowCount!=1: #skip header
            if not (row[12].value is None):
                if not (row[12].value in Mapping_RD_List) :
                    Mapping_RD_List.append(row[12].value)
                #else:
                    #print(row[12].value,"Duplicate entries in Excel")
    wb = None
    wb = load_workbook(filename = Excelfilepath, read_only=True)
    AgiliaReleaseSheet = wb["Agilia Release Data"]
    rowCount=0
    print("Processing Agilia Release Data Sheet")                
    for row in AgiliaReleaseSheet.rows:
        rowCount+=1
        if rowCount!=1: #skip header
            if not (row[0].value is None):
                #print(row[0].value)
                if not (row[0].value in Agilia_RD_List) :
                    Agilia_RD_List.append(row[0].value)
    wb = None
    
    print("Present in Mapping but not in Agilia Release Data")
    for id in Mapping_RD_List:
        if not(id in Agilia_RD_List) and len(id)>0:                
            print(id)
    
    print("Present in Agilia Release but not in Mapping")
    for id in Agilia_RD_List:
        if not(id in Mapping_RD_List) and len(id)>0:                
            print(id)
            
    Mapping_RD_List=[]
    Agilia_RD_List=[]
   # wb._archive.close()